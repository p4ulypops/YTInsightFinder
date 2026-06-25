#!/usr/bin/env python3
"""NuxTube export formats — turn an OmniFile into anything.

Each exporter takes an omni dict and writes one or more files to a target folder.
All formats derive from the same source-of-truth: omni.json.

Available formats
-----------------
  markdown     — Clean standalone Markdown note
  obsidian     — Obsidian vault note (YAML frontmatter, wikilinks, Dataview tags)
  csv          — Key points as CSV (one row per point)
  excel        — Multi-sheet XLSX (needs openpyxl; falls back to CSV)
  context_card — Compact 4-section LLM context card (C4 equivalent)
  hermes_skill — Hermes skill file (can be loaded with `hermes -z @skill`)
  llm_skill    — Generic LLM instruction document / system prompt

Usage
-----
    from nuxtube.exporters import export, FORMATS

    results = export(omni_dict, output_folder, formats=["obsidian", "csv"])
    # returns {format_name: output_path}

    # Or export everything:
    results = export(omni_dict, output_folder, formats=FORMATS)

CLI:
    python3 nuxtube.py --export ./youtube_videos/ai-agents/my-video --formats obsidian,csv,context_card
"""
import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

FORMATS = ["markdown", "obsidian", "csv", "excel", "context_card", "hermes_skill", "llm_skill"]


# ─── Helpers ────────────────────────────────────────────────────────────────

def _fmt_ts(ts) -> str:
    if ts is None:
        return ""
    t = int(float(ts))
    h, rem = divmod(t, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m}:{s:02d}"


def _meta(omni: dict) -> dict:
    return omni.get("metadata") or {}


def _kps(omni: dict) -> list:
    return (omni.get("key_points") or {}).get("key_points") or []


def _summary(omni: dict) -> str:
    return (omni.get("key_points") or {}).get("summary") or ""


def _chapters(omni: dict) -> list:
    pd = omni.get("player_data") or (_meta(omni).get("player_data") or {})
    return pd.get("chapters") or []


def _moments(omni: dict) -> list:
    return omni.get("synthesised_moments") or []


def _slug(s: str) -> str:
    import re
    s = re.sub(r"[^\w\s-]", "", s.lower())
    return re.sub(r"[\s-]+", "-", s).strip("-")[:60]


# ─── Markdown ────────────────────────────────────────────────────────────────

def export_markdown(omni: dict, out_dir: Path) -> str:
    """Clean standalone Markdown note."""
    m = _meta(omni)
    title = m.get("title") or "Untitled"
    channel = m.get("channel") or ""
    url = m.get("url") or (f"https://youtube.com/watch?v={m.get('video_id')}" if m.get("video_id") else "")
    dur = m.get("duration") or ""
    cat = m.get("category") or ""
    fetched = m.get("fetched_at", "")[:10]

    lines = [
        f"# {title}\n",
        f"> **Channel:** {channel}  ",
        f"> **Category:** {cat}  " if cat else "",
        f"> **Duration:** {dur}  " if dur else "",
        f"> **Archived:** {fetched}  " if fetched else "",
        f"> **Source:** [{url}]({url})" if url else "",
        "",
        "---",
        "",
    ]

    summary = _summary(omni)
    if summary:
        lines += ["## Summary\n", summary, ""]

    chapters = _chapters(omni)
    if chapters:
        lines += ["## Chapters\n"]
        for ch in chapters:
            ts = _fmt_ts(ch.get("start_time", 0))
            lines.append(f"- `{ts}` {ch.get('title', '')}")
        lines.append("")

    moments = _moments(omni)
    if moments:
        lines += ["## Key Moments (heatmap + chapters)\n"]
        for mo in moments:
            ts = _fmt_ts(mo["timestamp"])
            ch_title = mo.get("chapter_title") or ""
            score = f" · score {mo['heatmap_score']:.3f}" if mo.get("heatmap_score") else ""
            src = mo.get("source", "")
            badge = "🔥" if src == "heatmap" else "📑"
            excerpt = mo.get("transcript_excerpt") or ""
            lines.append(f"### {badge} `{ts}` {ch_title}{score}")
            if excerpt:
                lines.append(f"> {excerpt}")
            lines.append("")

    kps = _kps(omni)
    if kps:
        imp_icon = {"high": "⭐", "medium": "✅", "low": "ℹ️"}
        lines += ["## Key Points\n"]
        for kp in kps:
            ts = _fmt_ts(kp.get("timestamp")) if kp.get("timestamp") else ""
            ts_str = f" `⏱ {ts}`" if ts else ""
            imp = kp.get("importance", "medium")
            lines.append(f"### {imp_icon.get(imp,'✅')} {kp.get('title','')}{ts_str}")
            lines.append(f"**{kp.get('category','')}** · {imp}")
            lines.append("")
            lines.append(kp.get("lesson", ""))
            tags = kp.get("tags") or []
            if tags:
                lines.append("  " + " ".join(f"`{t}`" for t in tags))
            lines.append("")

    out_path = out_dir / "note.md"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    return str(out_path)


# ─── Obsidian ────────────────────────────────────────────────────────────────

def export_obsidian(omni: dict, out_dir: Path) -> str:
    """Obsidian-compatible note with YAML frontmatter, wikilinks, Dataview tags."""
    m = _meta(omni)
    title = m.get("title") or "Untitled"
    channel = m.get("channel") or ""
    url = m.get("url") or ""
    vid = m.get("video_id") or ""
    dur = m.get("duration") or ""
    cat = m.get("category") or "uncategorized"
    fetched = (m.get("fetched_at") or "")[:10] or datetime.now().strftime("%Y-%m-%d")
    thumb = m.get("thumbnail_url") or ""
    view_count = (omni.get("player_data") or {}).get("view_count")

    # Collect all tags from key points
    kp_tags: list = []
    for kp in _kps(omni):
        kp_tags.extend(kp.get("tags") or [])
    all_tags = sorted(set(kp_tags))

    # YAML frontmatter
    fm_lines = [
        "---",
        f'title: "{title.replace(chr(34), chr(39))}"',
        f"channel: \"[[{channel}]]\"",
        f"category: {cat}",
        f"duration: {dur}",
        f"archived: {fetched}",
        f"video_id: {vid}",
        f"source: youtube",
        f"url: {url}",
    ]
    if thumb:
        fm_lines.append(f"thumbnail: {thumb}")
    if view_count:
        fm_lines.append(f"view_count: {view_count}")
    fm_lines.append(f"tags: [{', '.join(all_tags)}]")
    fm_lines.append("---")
    fm_lines.append("")

    body = []
    body.append(f"# {title}\n")
    body.append(f"**Channel:** [[{channel}]] · **Category:** [[{cat}]] · **Duration:** {dur}")
    if url:
        body.append(f"**Watch:** [{url}]({url})")
    body.append("")

    summary = _summary(omni)
    if summary:
        body += ["## Summary\n", summary, ""]

    chapters = _chapters(omni)
    if chapters:
        body += ["## Chapters\n"]
        for ch in chapters:
            ts = _fmt_ts(ch.get("start_time", 0))
            ch_title = ch.get("title", "")
            # Wikilink the chapter title as a concept
            body.append(f"- `{ts}` [[{ch_title}]]")
        body.append("")

    moments = _moments(omni)
    if moments:
        body += ["## Key Moments\n"]
        body.append("| Timestamp | Chapter | Score | Excerpt |")
        body.append("|-----------|---------|-------|---------|")
        for mo in moments:
            ts = _fmt_ts(mo["timestamp"])
            ch_title = mo.get("chapter_title") or ""
            score = f"{mo['heatmap_score']:.3f}" if mo.get("heatmap_score") else "-"
            exc = (mo.get("transcript_excerpt") or "").replace("|", "·")[:80]
            body.append(f"| `{ts}` | {ch_title} | {score} | {exc} |")
        body.append("")

    kps = _kps(omni)
    if kps:
        imp_icon = {"high": "⭐", "medium": "✅", "low": "ℹ️"}
        body += ["## Key Points\n"]
        for kp in kps:
            ts = _fmt_ts(kp.get("timestamp")) if kp.get("timestamp") else ""
            ts_str = f" `{ts}`" if ts else ""
            imp = kp.get("importance", "medium")
            cat_kp = kp.get("category", "")
            body.append(f"### {imp_icon.get(imp,'✅')} {kp.get('title','')}{ts_str}")
            body.append(f"> [!info] {cat_kp} · {imp}")
            body.append(f"> {kp.get('lesson', '')}")
            tags = kp.get("tags") or []
            if tags:
                body.append("")
                body.append(" ".join(f"#{t.replace('-','_')}" for t in tags))
            body.append("")

    slug = _slug(title) or vid or "note"
    out_path = out_dir / f"{slug}.md"
    out_path.write_text("\n".join(fm_lines + body), encoding="utf-8")
    return str(out_path)


# ─── CSV ─────────────────────────────────────────────────────────────────────

def export_csv(omni: dict, out_dir: Path) -> str:
    """Key points as CSV — one row per point."""
    m = _meta(omni)
    title = m.get("title") or ""
    vid = m.get("video_id") or ""
    url = m.get("url") or ""

    fieldnames = ["video_id", "title", "url", "kp_id", "timestamp", "category",
                  "kp_title", "lesson", "importance", "tags"]

    out_path = out_dir / "key-points.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for kp in _kps(omni):
            writer.writerow({
                "video_id": vid,
                "title": title,
                "url": url,
                "kp_id": kp.get("id", ""),
                "timestamp": _fmt_ts(kp.get("timestamp")) if kp.get("timestamp") else "",
                "category": kp.get("category", ""),
                "kp_title": kp.get("title", ""),
                "lesson": kp.get("lesson", ""),
                "importance": kp.get("importance", ""),
                "tags": ";".join(kp.get("tags") or []),
            })
    return str(out_path)


def export_moments_csv(omni: dict, out_dir: Path) -> str:
    """Synthesised moments as CSV."""
    fieldnames = ["timestamp", "timestamp_fmt", "chapter_title", "heatmap_score", "source", "transcript_excerpt"]
    out_path = out_dir / "key-moments.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for mo in _moments(omni):
            writer.writerow({
                "timestamp": mo.get("timestamp", ""),
                "timestamp_fmt": _fmt_ts(mo.get("timestamp")),
                "chapter_title": mo.get("chapter_title", ""),
                "heatmap_score": mo.get("heatmap_score", ""),
                "source": mo.get("source", ""),
                "transcript_excerpt": mo.get("transcript_excerpt", ""),
            })
    return str(out_path)


# ─── Excel ───────────────────────────────────────────────────────────────────

def export_excel(omni: dict, out_dir: Path) -> str:
    """Multi-sheet XLSX. Requires openpyxl — falls back to CSV bundle if missing."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Graceful fallback: export a CSV bundle instead
        p1 = export_csv(omni, out_dir)
        p2 = export_moments_csv(omni, out_dir)
        return f"{p1},{p2}  (openpyxl not installed — saved as CSV)"

    m = _meta(omni)
    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="0D1117")
    HEADER_FONT = Font(bold=True, color="58A6FF")
    ALT_FILL = PatternFill("solid", fgColor="161B22")

    def style_header(row):
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=False)

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    style_header(ws[1])
    for k, v in [
        ("Title", m.get("title") or ""),
        ("Channel", m.get("channel") or ""),
        ("URL", m.get("url") or ""),
        ("Video ID", m.get("video_id") or ""),
        ("Category", m.get("category") or ""),
        ("Duration", m.get("duration") or ""),
        ("Archived", (m.get("fetched_at") or "")[:10]),
        ("Screenshots", m.get("media", {}).get("screenshot_count", "")),
        ("Clips", m.get("media", {}).get("clip_count", "")),
        ("Summary", _summary(omni)),
    ]:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80

    # Sheet 2: Key Points
    ws2 = wb.create_sheet("Key Points")
    ws2.append(["ID", "Timestamp", "Category", "Title", "Lesson", "Importance", "Tags"])
    style_header(ws2[1])
    for i, kp in enumerate(_kps(omni)):
        ts = _fmt_ts(kp.get("timestamp")) if kp.get("timestamp") else ""
        row = [
            kp.get("id", i + 1), ts, kp.get("category", ""),
            kp.get("title", ""), kp.get("lesson", ""),
            kp.get("importance", ""), ";".join(kp.get("tags") or []),
        ]
        ws2.append(row)
        if i % 2 == 0:
            for cell in ws2[ws2.max_row]:
                cell.fill = ALT_FILL
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 60

    # Sheet 3: Key Moments
    ws3 = wb.create_sheet("Key Moments")
    ws3.append(["Timestamp", "Chapter", "Heatmap Score", "Source", "Transcript Excerpt"])
    style_header(ws3[1])
    for i, mo in enumerate(_moments(omni)):
        ws3.append([
            _fmt_ts(mo.get("timestamp")),
            mo.get("chapter_title", ""),
            mo.get("heatmap_score") or "",
            mo.get("source", ""),
            mo.get("transcript_excerpt", ""),
        ])
        if i % 2 == 0:
            for cell in ws3[ws3.max_row]:
                cell.fill = ALT_FILL
    ws3.column_dimensions["E"].width = 80

    # Sheet 4: Chapters
    chapters = _chapters(omni)
    if chapters:
        ws4 = wb.create_sheet("Chapters")
        ws4.append(["Start Time", "Title"])
        style_header(ws4[1])
        for ch in chapters:
            ws4.append([_fmt_ts(ch.get("start_time", 0)), ch.get("title", "")])
        ws4.column_dimensions["B"].width = 60

    out_path = out_dir / "archive.xlsx"
    wb.save(str(out_path))
    return str(out_path)


# ─── Context Card (C4) ───────────────────────────────────────────────────────

def export_context_card(omni: dict, out_dir: Path) -> str:
    """Compact 4-section LLM context card.

    Format: WHO / WHAT / KEY INSIGHTS / KEY MOMENTS
    Drop this into any LLM's context window as background knowledge.
    Works beautifully as a system prompt addition or knowledge block.
    """
    m = _meta(omni)
    title = m.get("title") or "Untitled"
    channel = m.get("channel") or "Unknown"
    url = m.get("url") or ""
    dur = m.get("duration") or "?"
    cat = m.get("category") or ""
    fetched = (m.get("fetched_at") or "")[:10]
    views = (omni.get("player_data") or {}).get("view_count")

    lines = [
        f"# CONTEXT CARD: {title}",
        f"*Generated by NuxTube · {fetched}*",
        "",
        "---",
        "",
        "## 1. WHO & WHAT",
        f"- **Video:** {title}",
        f"- **Channel:** {channel}",
        f"- **Duration:** {dur}" + (f" · **Views:** {views:,}" if views else ""),
        f"- **Category:** {cat}" if cat else "",
        f"- **URL:** {url}" if url else "",
        "",
    ]

    summary = _summary(omni)
    if summary:
        lines += [
            "## 2. SUMMARY",
            summary,
            "",
        ]

    kps = _kps(omni)
    if kps:
        high = [k for k in kps if k.get("importance") == "high"]
        med = [k for k in kps if k.get("importance") == "medium"]
        show = (high + med)[:10] or kps[:10]

        lines += ["## 3. KEY INSIGHTS", ""]
        for kp in show:
            ts = _fmt_ts(kp.get("timestamp")) if kp.get("timestamp") else ""
            ts_str = f" [{ts}]" if ts else ""
            imp = "⭐" if kp.get("importance") == "high" else "→"
            lines.append(f"{imp} **{kp.get('title','')}**{ts_str}")
            lines.append(f"  {kp.get('lesson','')}")
            lines.append("")

    moments = _moments(omni)
    if moments:
        # Top 8 by heatmap score, or first 8 if no scores
        scored = [m for m in moments if m.get("heatmap_score")]
        unscored = [m for m in moments if not m.get("heatmap_score")]
        show_m = sorted(scored, key=lambda x: x["heatmap_score"], reverse=True)[:6] + unscored[:2]
        show_m = sorted(show_m, key=lambda x: x["timestamp"])

        lines += ["## 4. KEY MOMENTS", ""]
        for mo in show_m:
            ts = _fmt_ts(mo["timestamp"])
            ch = mo.get("chapter_title") or ""
            score_str = f" (engagement: {mo['heatmap_score']:.2f})" if mo.get("heatmap_score") else ""
            exc = mo.get("transcript_excerpt") or ""
            badge = "🔥" if mo.get("source") == "heatmap" else "📑"
            lines.append(f"{badge} **{ts}** — {ch}{score_str}")
            if exc:
                lines.append(f"  *\"{exc[:150]}\"*")
            lines.append("")

    lines += [
        "---",
        f"*Source: {url}*" if url else "",
    ]

    out_path = out_dir / "context_card.md"
    out_path.write_text("\n".join(l for l in lines if l is not None), encoding="utf-8")
    return str(out_path)


# ─── Hermes Skill ────────────────────────────────────────────────────────────

def export_hermes_skill(omni: dict, out_dir: Path) -> str:
    """Export as a Hermes skill file.

    Creates a markdown skill that can be loaded with `hermes -z @skill-name`.
    The skill encapsulates the video's knowledge as reusable expertise.
    """
    m = _meta(omni)
    title = m.get("title") or "Untitled"
    channel = m.get("channel") or "Unknown"
    url = m.get("url") or ""
    cat = m.get("category") or "general"
    fetched = (m.get("fetched_at") or "")[:10]

    kps = _kps(omni)
    summary = _summary(omni)

    slug = _slug(title) or "skill"
    skill_name = slug.replace("-", "_")

    # Derive a skill description from summary or title
    desc = summary[:120] if summary else f"Knowledge extracted from: {title}"

    lines = [
        f"# Skill: {skill_name}",
        "",
        "## Metadata",
        f"- **name:** {skill_name}",
        f"- **source:** YouTube — {title}",
        f"- **channel:** {channel}",
        f"- **category:** {cat}",
        f"- **archived:** {fetched}",
        f"- **url:** {url}" if url else "",
        "",
        "## Description",
        desc,
        "",
        "## System Context",
        "You are an expert assistant with deep knowledge extracted from the following video content.",
        "Apply these lessons when answering related questions. Be specific, cite timestamps when relevant.",
        "",
    ]

    if summary:
        lines += [
            "### Background",
            summary,
            "",
        ]

    if kps:
        lines += ["### Core Knowledge Base", ""]
        for kp in kps:
            ts = _fmt_ts(kp.get("timestamp")) if kp.get("timestamp") else ""
            ts_str = f" [{ts}]" if ts else ""
            imp = kp.get("importance", "medium")
            lines.append(f"**{kp.get('title','')}**{ts_str} ({imp} importance)")
            lines.append(kp.get("lesson", ""))
            tags = kp.get("tags") or []
            if tags:
                lines.append(f"Tags: {', '.join(tags)}")
            lines.append("")

    moments = _moments(omni)
    if moments:
        lines += ["### Key Moments Reference", ""]
        for mo in moments[:10]:
            ts = _fmt_ts(mo["timestamp"])
            ch = mo.get("chapter_title") or ""
            exc = mo.get("transcript_excerpt") or ""
            lines.append(f"- `{ts}` {ch}" + (f": {exc[:100]}" if exc else ""))
        lines.append("")

    lines += [
        "## Usage Examples",
        "",
        f"```",
        f"hermes -z @{skill_name} \"What are the key takeaways from this?\"",
        f"hermes -z @{skill_name} \"Summarise the most important moment\"",
        f"hermes -z @{skill_name} \"What tools or techniques were mentioned?\"",
        f"```",
        "",
        f"## Source",
        f"{url}" if url else f"Archived: {fetched}",
    ]

    out_path = out_dir / f"{slug}-skill.md"
    out_path.write_text("\n".join(l for l in lines if l is not None), encoding="utf-8")
    return str(out_path)


# ─── Generic LLM Skill / System Prompt ──────────────────────────────────────

def export_llm_skill(omni: dict, out_dir: Path) -> str:
    """Generic LLM skill — a structured system prompt + knowledge doc.

    Works with Claude, GPT, Gemini, local LLMs — any tool that accepts
    system context. Optimised for copy-paste into any AI interface.
    """
    m = _meta(omni)
    title = m.get("title") or "Untitled"
    channel = m.get("channel") or "Unknown"
    url = m.get("url") or ""
    cat = m.get("category") or "general"
    summary = _summary(omni)
    kps = _kps(omni)

    lines = [
        "<!-- LLM SKILL — copy everything below into a system prompt or knowledge block -->",
        "",
        "## ROLE & KNOWLEDGE BASE",
        "",
        f"You have been given expert knowledge extracted from a YouTube video:",
        f"**\"{title}\"** by {channel}",
        f"Category: {cat} | URL: {url}" if url else f"Category: {cat}",
        "",
    ]

    if summary:
        lines += [
            "**What this video covers:**",
            summary,
            "",
        ]

    high_kps = [k for k in kps if k.get("importance") == "high"]
    if high_kps:
        lines += ["**Most important insights (HIGH importance):**", ""]
        for kp in high_kps:
            ts = _fmt_ts(kp.get("timestamp")) if kp.get("timestamp") else ""
            ts_str = f" [{ts}]" if ts else ""
            lines.append(f"- **{kp.get('title','')}**{ts_str}: {kp.get('lesson','')}")
        lines.append("")

    med_kps = [k for k in kps if k.get("importance") != "high"]
    if med_kps:
        lines += ["**Additional insights:**", ""]
        for kp in med_kps:
            ts = _fmt_ts(kp.get("timestamp")) if kp.get("timestamp") else ""
            ts_str = f" [{ts}]" if ts else ""
            lines.append(f"- **{kp.get('title','')}**{ts_str}: {kp.get('lesson','')}")
        lines.append("")

    moments = _moments(omni)
    if moments:
        scored = sorted([mo for mo in moments if mo.get("heatmap_score")],
                        key=lambda x: x["heatmap_score"], reverse=True)[:5]
        if scored:
            lines += ["**Most-rewatched moments (viewer heatmap peaks):**", ""]
            for mo in sorted(scored, key=lambda x: x["timestamp"]):
                ts = _fmt_ts(mo["timestamp"])
                ch = mo.get("chapter_title") or ""
                exc = mo.get("transcript_excerpt") or ""
                lines.append(f"- `{ts}` {ch}: {exc[:120]}" if exc else f"- `{ts}` {ch}")
            lines.append("")

    # All key tags
    all_tags = sorted(set(t for kp in kps for t in (kp.get("tags") or [])))
    if all_tags:
        lines += [f"**Topic tags:** {', '.join(all_tags)}", ""]

    lines += [
        "## BEHAVIOUR INSTRUCTIONS",
        "",
        "When answering questions:",
        "1. Ground answers in the knowledge above — don't speculate beyond it",
        "2. Cite timestamps when referencing specific moments",
        "3. If asked about something not covered, say so clearly",
        "4. Be specific and actionable — prefer exact tool names, steps, techniques",
        "",
        "<!-- END LLM SKILL -->",
    ]

    out_path = out_dir / "llm_skill.md"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    return str(out_path)


# ─── Main export dispatcher ───────────────────────────────────────────────────

def export(omni: dict, output_folder: str, formats: List[str] = None) -> Dict[str, str]:
    """Export an OmniFile to one or more formats.

    Args:
        omni:           OmniFile dict (from build_omni() or loaded from omni.json)
        output_folder:  Where to write output files (usually the video archive folder)
        formats:        List of format names. Defaults to all formats.

    Returns:
        Dict mapping format name → output file path(s).
    """
    if formats is None:
        formats = FORMATS

    out_dir = Path(output_folder)
    out_dir.mkdir(parents=True, exist_ok=True)

    results = {}
    dispatch = {
        "markdown":     export_markdown,
        "obsidian":     export_obsidian,
        "csv":          export_csv,
        "excel":        export_excel,
        "context_card": export_context_card,
        "hermes_skill": export_hermes_skill,
        "llm_skill":    export_llm_skill,
    }

    for fmt in formats:
        fmt = fmt.strip().lower()
        fn = dispatch.get(fmt)
        if fn:
            try:
                results[fmt] = fn(omni, out_dir)
            except Exception as e:
                results[fmt] = f"ERROR: {e}"
        else:
            results[fmt] = f"ERROR: unknown format '{fmt}'"

    return results


def export_from_folder(folder: str, formats: List[str] = None) -> Dict[str, str]:
    """Convenience: load omni.json from folder and export."""
    from .omni import build_omni, write_omni
    folder_path = Path(folder)
    omni_path = folder_path / "omni.json"
    if not omni_path.exists():
        write_omni(folder)
    omni = build_omni(folder)
    if not omni:
        return {"error": "Could not build OmniFile"}
    return export(omni, folder, formats)
